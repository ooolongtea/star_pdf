import json
import os
import re
import sys
import torch
import cv2
import numpy as np
from PIL import Image
import subprocess
from typing import Optional, Union, List, Dict, Any, Tuple
from pathlib import Path
from huggingface_hub import hf_hub_download

# Model imports
from molscribe import MolScribe
from rxnscribe import MolDetect, RxnScribe

class MolCorefProcessor:
    """分子共指解析处理器"""
    SUPPORTED_EXTENSIONS = {'.png', '.jpg', '.jpeg'}
    
    def __init__(self, 
                 result_manager,
                 device: str = "cpu",
                 model_paths: Dict[str, str] = None):
        """
        初始化处理器
        
        参数:
            result_manager: 用于存储结果的ResultManager实例
            device: 计算设备('cpu'或'cuda:X')
            model_paths: 模型的可选路径(如果为None，则使用默认值)
        """
        self.result_manager = result_manager
        self.device = device
        
        self.model_paths = model_paths or {
            "molscribe": hf_hub_download('yujieq/MolScribe', 'swin_base_char_aux_1m.pth'),
            "molcoref": hf_hub_download("Ozymandias314/MolDetectCkpt", "coref_best_hf.ckpt"),
            "rxn": hf_hub_download("yujieq/RxnScribe", "pix2seq_reaction_full.ckpt")
        }
        
        # 初始化设备
        self._setup_device()
        
        # 加载模型
        self._load_models()
        
        # 标记模型加载状态
        self.models_loaded = True
    
    def _setup_device(self):
        """设置处理设备"""
        if self.device.startswith('cuda'):
            if not torch.cuda.is_available():
                raise RuntimeError("CUDA不可用！")
            self._device = torch.device(self.device)
    
    def _load_models(self):
        """加载所有所需的模型"""
        if not self.model_paths.get("molcoref"):
            default_path = os.path.expanduser("~/.cache/huggingface/hub/models--Ozymandias314--MolDetectCkpt/snapshots/7cf4ba5ffdae2aec35c8339693f1dafd70e0613c/coref_best_hf.ckpt")
            self.model_paths["molcoref"] = default_path
        
        try:
            # 加载分子检测模型
            self.model = MolDetect(
                self.model_paths["molcoref"],
                device=torch.device(self.device),
                coref=True  # 始终启用共指解析
            )
            
            # 加载分子结构识别模型
            self.molscribe = MolScribe(
                self.model_paths["molscribe"], 
                device=self.device
            )
            
            # 加载反应模型
            self.rxnmodel = RxnScribe(
                self.model_paths["rxn"],
                device=torch.device(self.device)
            )

            print("模型加载成功！---------------------------")
        except Exception as e:
            raise RuntimeError(f"加载模型失败：{str(e)}")
    
    def _should_skip(self, predictions: dict) -> bool:
        """检查是否需要跳过空结果"""
        bboxes = predictions.get('bboxes', [])
        corefs = predictions.get('corefs', [])
        return len(bboxes) == 0 and len(corefs) == 0
    
    def preprocess_image(self, image_path: Union[str, Path]) -> str:
        """
        图像预处理：去噪、锐化等
        
        参数:
            image_path: 原始图像路径
            
        返回:
            处理后的图像路径
        """
        img = cv2.imread(str(image_path), cv2.IMREAD_COLOR)
        if img is None:
            return str(image_path)  # 如果无法读取图像，返回原始路径
            
        # 转为灰度图
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        
        # 增强文本对比度
        binary = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
        
        # 去噪
        denoised = cv2.fastNlMeansDenoising(binary, h=10)
        
        # 锐化图像
        kernel = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]])
        sharpened = cv2.filter2D(denoised, -1, kernel)
        
        # 保存处理后的图像
        processed_dir = Path(image_path).parent / "processed"
        processed_dir.mkdir(exist_ok=True)
        
        processed_path = processed_dir / Path(image_path).name
        cv2.imwrite(str(processed_path), sharpened)
        
        return str(processed_path)
    
    def process_image(self, image_path: Union[str, Path], need_mol: bool = True, need_rxn: bool = True) -> Optional[str]:
        """
        处理单个图像文件，可独立控制是否进行分子和反应预测
        
        参数:
            image_path: 图像文件的路径
            need_mol: 是否进行分子预测
            need_rxn: 是否进行反应预测
            
        返回:
            如果处理成功则返回图像ID，否则返回None
        """
        image_path = Path(image_path).resolve()
        
        # 验证输入
        if not image_path.exists():
            print(f"输入文件不存在：{image_path}")
            return None
        
        if not self._is_supported_image(image_path):
            print(f"不支持的文件类型：{image_path.suffix}")
            return None
        
        # 图像预处理
        # processed_image_path = self.preprocess_image(image_path)
        processed_image_path = str(image_path)
        
        # 为此图像生成唯一ID
        image_id = f"{image_path.stem}"
        
        # 存储初始图像结果
        self.result_manager.store_image_result(image_id, str(image_path), {})
        
        # 执行预测（分子预测）
        if need_mol:
            try:
                predictions = self.model.predict_image_file(processed_image_path, coref=True)
                if not self._should_skip(predictions):
                    # 更新图像结果
                    self.result_manager.store_image_result(image_id, str(image_path), predictions)
                    # 处理分子和标识符
                    sheet_value = self._process_molecules_and_ids(predictions, image_path)
                    # 存储处理后的分子
                    for mol_data in sheet_value:
                        self.result_manager.store_molecule_result(image_id, mol_data)
            except Exception as e:
                print(f"分子预测失败: {e}")
        
        # 执行预测（反应预测）
        if need_rxn:
            try:
                rxn_predictions = self.rxnmodel.predict_image_file(processed_image_path, molscribe=True, ocr=True)
                # 处理反应数据
                reaction_data = self._process_reaction_data(rxn_predictions, image_path)
                # 存储处理后的反应数据
                for rxn_data in reaction_data:
                    self.result_manager.store_reaction_result(image_id, rxn_data)
                
                # 如果需要，生成并存储可视化
                if self.result_manager.output_config.get("visualization", True) and rxn_predictions:
                    vis_images = self.rxnmodel.draw_predictions(rxn_predictions, image_file=processed_image_path)
                    if vis_images:
                        self.result_manager.store_visualization(image_id, vis_images)
                    

                    
            except Exception as e:
                print(f"反应预测失败: {e}")
    
        return image_id
    
    def _is_supported_image(self, path: Path) -> bool:
        """检查文件类型是否受支持"""
        return path.suffix.lower() in self.SUPPORTED_EXTENSIONS
    
    def _process_reaction_data(self, rxn_predictions: list[dict], img_path: Path) -> List[dict]:
        """处理反应数据"""
        transformed_data = []
        
        # 检查rxn_predictions是否为空或无效
        if not rxn_predictions or not isinstance(rxn_predictions, list):
            return transformed_data
        
        # 处理每个反应
        for i, reaction in enumerate(rxn_predictions):
            # 提取反应物 SMILES
            reactants_smiles = ""
            if reaction.get("reactants") and len(reaction["reactants"]) > 0:
                reactants_smiles = ".".join(r.get("smiles", "") for r in reaction["reactants"] if r.get("smiles"))
            
            # 提取产物 SMILES
            product_smiles = ""
            if reaction.get("products") and len(reaction["products"]) > 0:
                product_smiles = ".".join(p.get("smiles", "") for p in reaction["products"] if p.get("smiles"))
            
            # 提取条件
            condition = ""
            if reaction.get("conditions"):
                condition_texts = []
                for c in reaction["conditions"]:
                    if "text" in c and c["text"]:
                        # text 可能是字符串或列表
                        if isinstance(c["text"], list):
                            condition_texts.extend(filter(None, c["text"]))
                        else:
                            condition_texts.append(c["text"])
                condition = ", ".join(condition_texts)
            
            # 创建新的反应数据字典
            new_reaction = {
                "reactants_smiles": reactants_smiles,
                "product_smiles": product_smiles,
                "product_coref": "",  # 默认为空，可以后续与分子结果进行匹配
                "condition": condition,
                "row_img": "",
                "image_path": str(img_path),  # 保存原图路径，用于在Excel中显示
                "reaction_id": i, # 用来对应可视化图
            }
            transformed_data.append(new_reaction)
        
        return transformed_data

    def _process_molecules_and_ids(self, predictions: dict, img_path: Path) -> List[dict]:
        """处理分子及其标识符"""
        original_image = Image.open(img_path)
        W, H = original_image.size
        results = []
        processed_mol_indices = set()  # 记录已处理的分子索引
        mol_id = 0
        
        # 处理共指解析（分子-标识符对）
        for coref_idx, (mol_idx, idt_idx) in enumerate(predictions.get('corefs', [])):
            # 处理分子
            mol_bbox = predictions['bboxes'][mol_idx]['bbox']
            mol_crop = self._crop_image(original_image, mol_bbox, W, H)
            smiles = self.process_mol(mol_crop)
            
            # 处理标识符
            idt_bbox = predictions['bboxes'][idt_idx]['bbox']
            idt_crop = self._crop_image(original_image, idt_bbox, W, H)
            coref = self.ocr_image(idt_crop)
            
            results.append({
                "id": mol_id,
                "compound_smiles": smiles,
                "coref": coref,
                "compound_name": "",  # 添加空字段以符合Excel的格式
                "mol_idx": mol_idx,
                "idt_idx": idt_idx
            })
            processed_mol_indices.add(mol_idx)  # 记录已处理的分子
            mol_id += 1
        
        # 处理未被共指解析覆盖的[Mol]
        for mol_idx, bbox in enumerate(predictions['bboxes']):
            if bbox['category'] == '[Mol]' and mol_idx not in processed_mol_indices:
                mol_bbox = bbox['bbox']
                mol_crop = self._crop_image(original_image, mol_bbox, W, H)
                smiles = self.process_mol(mol_crop)
                
                results.append({
                    "id": mol_id,    
                    "compound_smiles": smiles,
                    "coref": "",  # 未关联的标识符设为空
                    "compound_name": ""  # 添加空字段以符合Excel的格式
                })
                processed_mol_indices.add(mol_idx)
                mol_id += 1
        
        return results
    
    def _crop_image(self, image: Image, bbox: List[float], W: int, H: int) -> Image:
        """使用归一化坐标裁剪图像"""
        x_1 = bbox[0] * W
        y_1 = bbox[1] * H
        x_2 = min(W, bbox[2] * W)
        y_2 = min(H, bbox[3] * H)
        return image.crop((x_1, y_1, x_2, y_2))
    
    def process_mol(self, image: Image) -> str:
        """读取分子结构"""
        image_np = np.array(image)
        results = self.molscribe.predict_image(image_np)
        return results['smiles'] if results else ""
    
    def ocr_image(self, image: Image) -> str:
        """使用外部进程或直接运行图像OCR"""
        # 保存临时图像文件
        temp_image_path = "/tmp/temp_image.png"
        image.save(temp_image_path)
        
        # 获取当前脚本目录的路径
        script_dir = Path(__file__).parent
        script_path = script_dir / "ocr_runner.py"
        
        # 使用子进程调用PaddleOCR环境
        try:
            # 尝试找到安装了PaddleOCR的Python可执行文件
            python_paths = [
                os.environ.get("PADDLEOCR_PYTHON", ""),  # 如果设置了，使用环境变量
                os.path.expanduser("~/.conda/envs/zxy-paddleocr/bin/python"),  # 默认路径
                "python"  # 退回到系统Python
            ]
            
            # 使用第一个可用的Python路径
            python_executable = next((p for p in python_paths if p and os.path.exists(p)), "python")
            
            # 运行子进程
            result = subprocess.run(
                [python_executable, str(script_path), temp_image_path],
                capture_output=True,
                text=True,
                check=True,
                env={"PYTHONWARNINGFILTER": "ignore"}  # 忽略警告
            )
            
            # 使用正则表达式提取OCR结果
            pattern = r'paddleocr--------------:\s*\[([^\]]+)\]'
            match = re.search(pattern, result.stdout)
            
            if match:
                # 解析列表内容
                content = match.group(1).strip()
                if not content:  # 处理空列表
                    return ""
                
                items = [s.strip().strip("'") for s in content.split(",")]
                return items[0] if items else ""
            return ""
        
        except subprocess.CalledProcessError as e:
            print(f"OCR子进程失败：{e.stderr}", file=sys.stderr)
            return ""

    def write_to_excel(self, output_file: Union[str, Path], molecules_data: List[dict] = None, reactions_data: List[dict] = None):
        """
        将数据写入Excel文件
        
        参数:
            output_file: Excel文件路径
            molecules_data: 分子数据列表
            reactions_data: 反应数据列表
        """
        import openpyxl
        from openpyxl.drawing.image import Image
        from openpyxl.styles import Alignment, PatternFill
        import os
        
        output_file = Path(output_file)
        
        # 使用存储在结果管理器中的数据（如果没有提供）
        if molecules_data is None:
            molecules_data = []
            for image_mols in self.result_manager.get_all_molecule_results().values():
                molecules_data.extend(image_mols)
                
        if reactions_data is None:
            reactions_data = []
            for image_rxns in self.result_manager.get_all_reaction_results().values():
                reactions_data.extend(image_rxns)
        
        try:
            # 尝试打开现有的工作簿
            wb = openpyxl.load_workbook(output_file)
        except FileNotFoundError:
            # 如果文件不存在，创建新的工作簿
            wb = openpyxl.Workbook()
            # 保留默认的Sheet，只有在确定有其他sheet后才移除
        
        # 确保至少有一个可见的sheet
        has_visible_sheets = False
        
        # 处理Sheet2（反应数据）
        if reactions_data:
            has_visible_sheets = True
            sheet_name = "Sheet2"
            headers = ["id", "reactants_smiles", "product_smiles", "product_coref", "condition", "row_img"]
            
            # 如果sheet不存在，创建新的sheet
            if sheet_name not in wb.sheetnames:
                sheet = wb.create_sheet(sheet_name)
                # 添加表头
                sheet.append(headers)
                # 设置表头样式
                header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                for col_num, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                sheet = wb[sheet_name]
            
            # 获取下一个id值 
            next_id = 1
            if len(sheet['A']) > 1:  # 如果已有数据（表头+至少1行数据）
                next_id = len(sheet['A'])  # 假设id存放在A列
            

            # 写入反应数据
            for reaction in reactions_data:
                # 设置id
                reaction["id"] = next_id
                
                # 准备行数据
                row_data = {
                    "id": next_id,
                    "reactants_smiles": reaction.get("reactants_smiles", ""),
                    "product_smiles": reaction.get("product_smiles", ""),
                    "product_coref": reaction.get("product_coref", ""),
                    "condition": reaction.get("condition", ""),
                    "row_img": ""  # 置空，直接插入图片
                }
                
                # 写入数据行
                row = [row_data.get(header, "") for header in headers]
                sheet.append(row)
                
                # 获取当前行号
                current_row = sheet.max_row
                
                # 设置文本单元格的对齐方式
                for col_num in range(1, len(headers)):
                    cell = sheet.cell(row=current_row, column=col_num)
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # 插入图片
                image_path = reaction.get("image_path", "")
                # 如果可视化，用处理后图片
                if self.result_manager.output_config.get("visualization", True):
                    image_id = reaction.get("reaction_id", "")
                    parent_dir = Path(image_path).parent.parent
                    new_dir = parent_dir / "image_visualizations"
                    file_stem = Path(image_path).stem
                    new_filename = f"{file_stem}_visualization_{image_id}.png"
                    image_path = new_dir / new_filename


                if image_path and os.path.exists(image_path):
                    try:
                        # 设置图片大小和缩放比例
                        img = Image(image_path)
                        # 根据图片实际尺寸调整大小，最大宽度100像素
                        max_width = 100
                        if img.width > max_width:
                            scale_ratio = max_width / img.width
                            img.width = max_width
                            img.height = int(img.height * scale_ratio)
                        
                        # 添加图片到row_img列
                        img_col = headers.index("row_img") + 1  # 列索引从1开始
                        img_cell = f"{chr(64 + img_col)}{current_row}"  # 例如 F2
                        
                        # 设置行高以适应图片
                        row_height = max(60, img.height * 0.75)  # 确保至少有60的行高
                        sheet.row_dimensions[current_row].height = row_height
                        
                        # 设置列宽以适应图片
                        col_letter = chr(64 + img_col)
                        sheet.column_dimensions[col_letter].width = max(15, img.width * 0.14)
                        
                        # 添加图片到单元格，并相对于单元格左上角进行偏移
                        img.anchor = img_cell
                        
                        # 添加图片到工作表
                        sheet.add_image(img)
                        
                        # 设置包含图片的单元格对齐方式
                        cell = sheet.cell(row=current_row, column=img_col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                    except Exception as e:
                        print(f"无法插入图片 {image_path}: {e}")
                        # 图片插入失败，在单元格中写入路径
                        sheet.cell(row=current_row, column=headers.index("row_img") + 1).value = image_path
                else:
                    # 如果图片路径不存在，在单元格中写入路径
                    if image_path:
                        sheet.cell(row=current_row, column=headers.index("row_img") + 1).value = image_path
                
                next_id += 1
        
        # 处理Sheet3（分子数据）
        if molecules_data:
            has_visible_sheets = True
            sheet_name = "Sheet3"
            headers = ["id", "coref", "compound_name", "compound_smiles"]
            
            # 如果sheet不存在，创建新的sheet
            if sheet_name not in wb.sheetnames:
                sheet = wb.create_sheet(sheet_name)
                # 添加表头
                sheet.append(headers)
                # 设置表头样式
                header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                for col_num, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                sheet = wb[sheet_name]
            
            # 获取下一个id值
            next_id = 1
            if len(sheet['A']) > 1:  # 如果已有数据（表头+至少1行数据）
                next_id = len(sheet['A'])  # 假设id存放在A列
            
            # 写入分子数据
            for molecule in molecules_data:
                # 设置id
                molecule["id"] = next_id
                
                # 写入数据
                row = [next_id, 
                    molecule.get("coref", ""), 
                    molecule.get("compound_name", ""), 
                    molecule.get("compound_smiles", "")]
                sheet.append(row)
                
                # 获取当前行号
                current_row = sheet.max_row
                
                # 设置单元格对齐方式
                for col_num in range(1, len(headers) + 1):
                    cell = sheet.cell(row=current_row, column=col_num)
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                next_id += 1
        
        # 如果没有反应数据和分子数据，确保至少有一个sheet
        if not has_visible_sheets:
            # 检查是否存在默认sheet
            default_sheet_exists = "Sheet" in wb.sheetnames
            if not default_sheet_exists:
                # 如果没有默认sheet，创建一个新的sheet
                wb.create_sheet("Sheet1")
        else:
            # 如果有其他可见sheet，并且默认的"Sheet"还存在，则可以移除它
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                wb.remove(wb["Sheet"])
        
        # 保存工作簿
        wb.save(output_file)
        return output_file