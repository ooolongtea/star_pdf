import os
from pathlib import Path
import json
from typing import Dict, List, Optional, Union, Any
import numpy as np
from PIL import Image
import openpyxl
from openpyxl.drawing.image import Image as XlImage
import gc

class ResultManager:
    """
    中央类，用于在处理过程中管理中间结果和最终结果。
    """

    def __init__(self, output_config: Dict[str, bool] = None):
        """
        使用控制要生成的输出的配置初始化ResultManager。

        参数:
            output_config: 控制要生成哪些输出的字典
                - json_results: 保存JSON结果
                - excel_results: 保存Excel结果
                - visualization: 保存可视化图像
                - intermediate_files: 保存中间处理文件
        """
        self.output_config = output_config or {
            "json_results": True,
            "excel_results": True,
            "visualization": True,
            "intermediate_files": False # 每个图片的处理后json
        }

        # 各种结果的存储
        self.patent_results = {}  # 专利级结果
        self.image_results = {}   # 图像级结果
        self.molecule_results = {}  # 分子结果（Sheet3数据）
        self.reaction_results = {}  # 反应结果（Sheet2数据）
        self.experiment_results = {}  # 实验结果（Sheet1数据）
        self.visualization_data = {}  # 可视化数据

        self.molscribe_results = {}  # molscribe检测结果



    def store_molecule_result(self, image_id: str, molecule_data: Dict[str, Any]):
        """存储分子检测结果（用于Sheet3）"""
        if image_id not in self.molecule_results:
            self.molecule_results[image_id] = []

        self.molecule_results[image_id].append(molecule_data)
        return molecule_data

    def store_reaction_result(self, image_id: str, reaction_data: Dict[str, Any]):
        """存储反应数据（用于Sheet2）"""
        if image_id not in self.reaction_results:
            self.reaction_results[image_id] = []

        self.reaction_results[image_id].append(reaction_data)
        return reaction_data

    def store_experiment_result(self, image_id: str, experiment_data: Dict[str, Any]):
        """存储实验数据（用于Sheet1）"""
        if image_id not in self.experiment_results:
            self.experiment_results[image_id] = []

        self.experiment_results[image_id].append(experiment_data)
        return experiment_data

    def store_image_result(self, image_id: str, image_path: str, predictions: Dict[str, Any]):
        """存储图像检测结果"""
        self.image_results[image_id] = {
            "image_path": image_path,
            "predictions": predictions
        }
        return self.image_results[image_id]

    def store_patent_result(self, patent_id: str, data: Dict[str, Any]):
        """存储专利级结果"""
        # 保留processed_images字段
        if patent_id in self.patent_results:
            # 如果专利已存在，确保保留processed_images字段
            if "processed_images" in data:
                existing_images = set(self.patent_results[patent_id].get("processed_images", []))
                new_images = set(data.get("processed_images", []))
                # 合并图像列表
                all_images = list(existing_images.union(new_images))
                data["processed_images"] = all_images
            else:
                # 如果新数据没有processed_images字段，保留旧的
                data["processed_images"] = self.patent_results[patent_id].get("processed_images", [])

        self.patent_results[patent_id] = data
        return self.patent_results[patent_id]

    def store_molscribe_result(self, image_id: str, molscribe_data: Dict[str, Any]):
        """存储molscribe检测结果"""
        if image_id not in self.molscribe_results:
            self.molscribe_results[image_id] = []

        self.molscribe_results[image_id].append(molscribe_data)
        return molscribe_data


    def store_visualization(self, image_id: str, visualization_data: np.ndarray):
        """存储可视化数据"""
        if self.output_config.get("visualization", True):
            self.visualization_data[image_id] = visualization_data

    def get_molecule_results(self, image_id: str) -> List[Dict[str, Any]]:
        """获取特定图像的分子结果"""
        return self.molecule_results.get(image_id, [])

    def get_reaction_results(self, image_id: str) -> List[Dict[str, Any]]:
        """获取特定图像的反应结果"""
        return self.reaction_results.get(image_id, [])

    def get_experiment_results(self, image_id: str) -> List[Dict[str, Any]]:
        """获取特定图像的实验结果"""
        return self.experiment_results.get(image_id, [])

    def get_molscribe_results(self, image_id: str) -> List[Dict[str, Any]]:
        """获取特定图像的molscribe结果"""
        return self.molscribe_results.get(image_id, [])

    def get_all_molecule_results(self) -> Dict[str, List[Dict[str, Any]]]:
        """获取所有分子结果"""
        return self.molecule_results

    def get_all_reaction_results(self) -> Dict[str, List[Dict[str, Any]]]:
        """获取所有反应结果"""
        return self.reaction_results

    def get_all_experiment_results(self) -> Dict[str, List[Dict[str, Any]]]:
        """获取所有实验结果"""
        return self.experiment_results

    def get_image_result(self, image_id: str) -> Dict[str, Any]:
        """获取图像结果"""
        return self.image_results.get(image_id, {})

    def get_all_image_results(self) -> Dict[str, Dict[str, Any]]:
        """获取所有图像结果"""
        return self.image_results

    def get_patent_result(self, patent_id: str) -> Dict[str, Any]:
        """获取专利结果"""
        return self.patent_results.get(patent_id, {})

    def get_all_patent_results(self) -> Dict[str, Dict[str, Any]]:
        """获取所有专利结果"""
        return self.patent_results

    def get_visualization(self, image_id: str) -> Optional[np.ndarray]:
        """获取可视化数据"""
        return self.visualization_data.get(image_id)

    def save_results(self, output_dir: Path):
        """
        根据配置保存所有结果

        参数:
            output_dir: 保存结果的目录
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        # 如果配置了，保存JSON结果
        if self.output_config.get("json_results", True):
            self._save_json_results(output_dir)

        # 如果配置了，保存可视化
        if self.output_config.get("visualization", True):
            # 确保可视化图像位于主输出目录的子目录中
            vis_dir = output_dir / "image_visualizations"
            self._save_visualizations(vis_dir)

        # 返回所有保存文件的路径
        return {
            "output_dir": str(output_dir)
        }

    def _save_json_results(self, output_dir: Path):
        """保存JSON结果"""
        # 保存专利结果（按图像分类）
        for patent_id, patent_data in self.patent_results.items():
            # 重组数据结构，按图像分类结果
            images_data = {}

            # 获取所有与该专利相关的图像ID
            image_ids = patent_data.get("processed_images", [])

            # 为每个图像收集相关结果
            for image_id in image_ids:
                molecules = self.get_molecule_results(image_id)
                reactions = self.get_reaction_results(image_id)
                # experiments = self.get_experiment_results(image_id)

                # 存储该图像的所有相关数据
                images_data[image_id] = {
                    "molecules": molecules,
                    "reactions": reactions,
                    # "experiments": experiments,
                    "image_path": self.get_image_result(image_id).get("image_path", "")
                }

            # 创建新的专利数据结构
            organized_patent_data = {
                "patent_id": patent_id,
                "total_images": len(image_ids),
                "images": images_data
            }

            # 保存整理后的JSON结果到主输出目录到主输出目录
            patent_file = output_dir / f"{patent_id}_results.json"
            with open(patent_file, "w", encoding="utf-8") as f:
                json.dump(organized_patent_data, f, ensure_ascii=False, indent=2)

        # 始终保存单独的图像结果，不再依赖intermediate_files选项
        # 在主输出目录下创建子目录
        image_results_dir = output_dir / "image_results"
        image_results_dir.mkdir(exist_ok=True)

        for image_id, image_data in self.image_results.items():
            # 为该图像收集所有结果
            molecules = self.get_molecule_results(image_id)
            reactions = self.get_reaction_results(image_id)
            # experiments = self.get_experiment_results(image_id)

            # 创建完整的图像数据
            complete_image_data = {
                "image_id": image_id,
                "image_path": image_data.get("image_path", ""),
                "molecules": molecules,
                "reactions": reactions,
                # "experiments": experiments,
                "predictions": image_data.get("predictions", {})
            }

            # 保存单独的图像结果到子目录中
            image_file = image_results_dir / f"{image_id}_image_results.json"
            with open(image_file, "w", encoding="utf-8") as f:
                json.dump(complete_image_data, f, ensure_ascii=False, indent=2)

    def _save_visualizations(self, output_dir: Path):
        """保存可视化图像"""
        output_dir.mkdir(parents=True, exist_ok=True)
        for image_id, vis_data in self.visualization_data.items():
            for i, vis_img in enumerate(vis_data):
                # 保存每个可视化图像
                if isinstance(vis_img, np.ndarray):
                    vis_file = output_dir / f"{image_id}_visualization_{i}.png"
                    img = Image.fromarray(vis_img)
                    img.save(vis_file)
                else:
                    print(f"Warning: 可视化数据不是有效的图像格式，跳过保存：{image_id}")

    # 已废弃
    def to_excel(self, output_file: Path):
        """
        将结果导出到Excel

        参数:
            output_file: Excel文件的路径
        """
        if not self.output_config.get("excel_results", True):
            return None

        try:
            # 尝试打开现有的工作簿
            wb = openpyxl.load_workbook(output_file)
        except FileNotFoundError:
            # 如果文件不存在，创建新的工作簿
            wb = openpyxl.Workbook()
            # 移除默认创建的sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        # # 处理Sheet1（实验数据）
        # self._write_to_sheet1(wb, output_file)

        # # 处理Sheet2（反应数据）
        self._write_to_sheet2(wb, output_file)

        # 处理Sheet3（分子数据）
        self._write_to_sheet3(wb, output_file)

        # 保存工作簿
        wb.save(output_file)
        return output_file

    def _write_to_sheet1(self, workbook, output_file):
        """写入Sheet1（实验数据）"""
        sheet_name = "Sheet1"
        headers = ["id", "reactants_name", "reactants_smiles", "product_name", "product_smiles",
                  "temperature", "time", "yield", "condition", "experiment_manipulation", "row_text"]

        # 如果sheet不存在，创建新的sheet
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            # 添加表头
            sheet.append(headers)
        else:
            sheet = workbook[sheet_name]

        # 获取下一个id值
        next_id = len(sheet['A'])  # 假设id存放在A列

        # 写入实验数据
        for image_id, experiments in self.experiment_results.items():
            for experiment in experiments:
                # 设置id
                experiment["id"] = next_id

                # 写入数据
                row = [experiment.get(header, None) for header in headers]
                sheet.append(row)

                next_id += 1

    def _write_to_sheet2(self, workbook, output_file):
        """写入Sheet2（反应数据）"""
        sheet_name = "Sheet2"
        headers = ["id", "reactants_smiles", "product_smiles", "product_coref", "condition", "row_img"]

        # 如果sheet不存在，创建新的sheet
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            # 添加表头
            sheet.append(headers)
        else:
            sheet = workbook[sheet_name]

        # 获取下一个id值
        next_id = 1
        if len(sheet['A']) > 1:  # 如果已有数据（表头+至少1行数据）
            next_id = len(sheet['A'])  # 假设id存放在A列

        # 写入反应数据
        for image_id, reactions in self.reaction_results.items():
            for reaction in reactions:
                # 设置id
                reaction["id"] = next_id

                # 准备行数据
                row_data = {
                    "id": next_id,
                    "reactants_smiles": reaction.get("reactants_smiles", ""),
                    "product_smiles": reaction.get("product_smiles", ""),
                    "product_coref": reaction.get("product_coref", ""),
                    "condition": reaction.get("condition", ""),
                    "row_img": reaction.get("image_path","")  # 插入图像有问题，先用地址代替
                }

                # 写入数据行
                row = [row_data.get(header, "") for header in headers]
                sheet.append(row)

                # # 如果有图片路径，插入图片
                # image_path = reaction.get("image_path")
                # if image_path:
                #     try:
                #         # 调整图像大小以适合单元格
                #         img = Image.open(image_path)
                #         max_height = 120  # 最大高度（像素）
                #         max_width = 160   # 最大宽度（像素）
                #         img.thumbnail((max_width, max_height))

                #         # 保存临时调整大小的图像
                #         temp_path = output_file.parent / f"temp_{next_id}.png"
                #         img.save(temp_path)

                #         # 添加到Excel中
                #         xl_img = XlImage(str(temp_path))
                #         cell_address = f"F{next_id + 1}"  # row_img

                #         # 调整行高以适应图像
                #         row_number = next_id + 1
                #         current_height = sheet.row_dimensions[row_number].height or 0
                #         new_height = max(75, current_height)
                #         sheet.row_dimensions[row_number].height = new_height

                #         # 添加图片
                #         sheet.add_image(xl_img, cell_address)

                #         # 删除临时文件
                #         # os.remove(temp_path)
                #     except Exception as e:
                #         print(f"插入图片失败：{e}")

                next_id += 1


    def _write_to_sheet3(self, workbook, output_file):
        """写入Sheet3（分子数据）"""
        sheet_name = "Sheet3"
        headers = ["id", "coref", "compound_name", "compound_smiles"]

        # 如果sheet不存在，创建新的sheet
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            # 添加表头
            sheet.append(headers)
        else:
            sheet = workbook[sheet_name]

        # 获取下一个id值
        next_id = len(sheet['A'])  # 假设id存放在A列

        # 写入分子数据
        for image_id, molecules in self.molecule_results.items():
            for molecule in molecules:
                # 设置id
                molecule["id"] = next_id

                # 写入数据
                row = [molecule.get(header, None) for header in headers]
                sheet.append(row)

                next_id += 1


    def save_molcoref_results(self, image_id:str, output_dir: Path):
        """
        保存molcoref解析结果，单个图像使用暴露的接口(测试用的，不用管)

        参数:
            output_dir: 保存结果的目录
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        saved_files = {"json_files": [], "visualization_files": []}

        # 保存JSON结果
        if self.output_config.get("json_results", True):
            # 处理单个图像的结果
            json_file = output_dir / f"{image_id}_results.json"
            molecules = self.get_molecule_results(image_id)
            reaction = self.get_reaction_results(image_id)

            # 创建结果字典
            result_data = {
                "image_id": image_id,
                "molecules_count": len(molecules),
                "molecules": molecules,
                "reaction_count": len(reaction),
                "reaction": reaction,
                "image_path": self.get_image_result(image_id).get("image_path", ""),
            }

            # 保存JSON
            with open(json_file, "w", encoding="utf-8") as f:
                json.dump(result_data, f, ensure_ascii=False, indent=2)

            saved_files["json_files"].append(str(json_file))

        # 保存可视化
        if self.output_config.get("visualization", True):
            vis_data = self.get_visualization(image_id)
            if vis_data is not None:
                vis_file = output_dir / f"{image_id}_visualization.png"
                img = Image.fromarray(vis_data)
                img.save(vis_file)
                saved_files["visualization_files"].append(str(vis_file))

        return saved_files

    def save_molscribe_results(self, image_id:str, output_dir: Path):
        """
        保存molscribe检测结果，单个图像使用暴露的接口
        参数:
            output_dir: 保存结果的目录
        保存为json
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        # 如果配置了，保存json结果
        if self.output_config.get("json_results", True):
            # 处理单个图像的结果
            json_file = output_dir / f"{image_id}_molscribe_results.json"
            molscribe_data = self.get_molscribe_results(image_id)

            # 创建结果字典
            result_data = {
                "molscribe_data": molscribe_data,
                "image_path": self.get_image_result(image_id).get("image_path", ""),
            }

            # 保存JSON
            with open(json_file, "w", encoding="utf-8") as f:
                json.dump(result_data, f, ensure_ascii=False, indent=2)

        return str(json_file)

    def save_rxnscribe_results(self, image_id:str, output_dir: Path):
        """
        保存rxnscribe检测结果，单个图像使用暴露的接口
        保存为json，可视化图片
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        # 如果配置了，保存json结果
        if self.output_config.get("json_results", True):
            # 处理单个图像的结果
            json_file = output_dir / f"{image_id}_rxnscribe_results.json"
            rxnscribe_data = self.get_reaction_results(image_id)

            # 创建结果字典
            result_data = {
                "rxnscribe_data": rxnscribe_data,
                "image_path": self.get_image_result(image_id).get("image_path", ""),
            }

            # 保存JSON
            with open(json_file, "w", encoding="utf-8") as f:
                json.dump(result_data, f, ensure_ascii=False, indent=2)

        # 如果配置了，保存可视化
        if self.output_config.get("visualization", True):
            self._save_visualizations(output_dir)

        return str(output_dir)

    def clear_results(self):
        """清理所有结果数据，释放内存"""
        # 清空各种存储
        self.patent_results.clear()
        self.image_results.clear()
        self.molecule_results.clear()
        self.reaction_results.clear()
        self.experiment_results.clear()
        self.visualization_data.clear()
        self.molscribe_results.clear()

        # 进行垃圾回收
        gc.collect()
        
        # 进行垃圾回收
        gc.collect()

        # 返回自身以支持链式调用
        return self


