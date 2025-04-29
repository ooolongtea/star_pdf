import os
import logging
import shutil
from pathlib import Path
import tempfile
import openpyxl
from openpyxl.utils import get_column_letter
from typing import List, Optional, Dict, Any
import subprocess
from base_parser import DocumentParser

# 配置日志
logger = logging.getLogger("excel_parser")

class ExcelParser(DocumentParser):
    """本地Excel文档解析器，支持.xls和.xlsx格式"""
    
    def parse(self, file_path: str, output_dir: str) -> Optional[str]:
        """
        解析Excel（支持.xls和.xlsx），转换为Markdown，并存放到结构化目录
        
        Args:
            file_path: Excel文件路径
            output_dir: 输出目录路径
            
        Returns:
            str: 生成的Markdown文件路径，失败时返回None
        """
        try:
            # 验证输入文件
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"输入文件不存在: {file_path}")
                
            # 创建输出目录
            os.makedirs(output_dir, exist_ok=True)
            
            # 获取文件名（不含扩展名）
            file_name = os.path.splitext(os.path.basename(file_path))[0]
            
            # 统一处理为.xlsx格式
            xlsx_path = self._convert_to_xlsx_if_needed(file_path, output_dir)
            
            # 处理xlsx文件
            md_path = self._parse_xlsx(xlsx_path, output_dir, file_name)
            
            # 复制原始文件到输出目录
            orig_file_ext = os.path.splitext(file_path)[1]
            orig_file_dest = os.path.join(output_dir, f"{file_name}{orig_file_ext}")
            shutil.copy2(file_path, orig_file_dest)
            logger.info(f"已复制原始文件: {orig_file_dest}")
            
            # 清理临时文件
            if xlsx_path != file_path and os.path.exists(xlsx_path):
                os.remove(xlsx_path)
                logger.debug(f"已删除临时文件: {xlsx_path}")
                
            return md_path
                
        except Exception as e:
            logger.error(f"Excel解析失败: {str(e)}")
            return None

    def _convert_to_xlsx_if_needed(self, file_path: str, output_dir: str) -> str:
        """将.xls转换为.xlsx（如需要）"""
        if file_path.lower().endswith('.xlsx'):
            return file_path
            
        # 创建临时目录用于转换
        temp_dir = tempfile.mkdtemp()
        try:
            output_path = os.path.join(temp_dir, os.path.basename(file_path).split(".")[0] + '.xlsx')
            
            try:
                subprocess.run([
                    "soffice", "--headless", "--convert-to", "xlsx",
                    "--outdir", temp_dir, file_path
                ], check=True, timeout=30)
                logger.info(f"已转换.xls到.xlsx: {output_path}")
                return output_path
            except subprocess.TimeoutExpired:
                raise RuntimeError("文档转换超时，请确保已安装LibreOffice")
            except Exception as e:
                raise RuntimeError(f"文档转换失败: {str(e)}")
        except Exception as e:
            logger.error(f"转换文档失败: {str(e)}")
            shutil.rmtree(temp_dir, ignore_errors=True)
            raise
        
    def _parse_xlsx(self, file_path: str, output_dir: str, file_name: str) -> str:
        """解析.xlsx并转换为Markdown"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            all_md_lines = []
            
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                md_lines = [f"\n## {sheet_name}\n\n"]
                
                # 处理表头
                headers = self._get_headers(ws)
                md_lines.append(self._create_markdown_table_row(headers))
                md_lines.append(self._create_markdown_separator(len(headers)))
                
                # 处理数据行（包括合并单元格）
                merged_ranges = ws.merged_cells.ranges
                for row in ws.iter_rows(min_row=2):
                    row_data = self._process_row(row, merged_ranges)
                    md_lines.append(self._create_markdown_table_row(row_data))
                
                all_md_lines.extend(md_lines)
            
            # 保存Markdown文件
            md_path = os.path.join(output_dir, f"{file_name}.md")
            
            with open(md_path, 'w', encoding='utf-8') as f:
                f.writelines(all_md_lines)
                
            logger.info(f"Excel解析完成: {md_path}")
            return md_path
            
        except Exception as e:
            raise RuntimeError(f"解析XLSX文件失败: {str(e)}")

    # 辅助方法 ------------------------------------------------------
    
    def _get_headers(self, worksheet) -> List[str]:
        """获取表头行"""
        return [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
        ]
    
    def _process_row(self, row, merged_ranges) -> List[str]:
        """处理数据行，处理合并单元格"""
        row_data = []
        for cell in row:
            cell_value = self._get_cell_value(cell, merged_ranges)
            row_data.append(str(cell_value).replace('\n', '<br>'))  # 处理换行符
        return row_data
    
    def _get_cell_value(self, cell, merged_ranges):
        """获取单元格值，处理合并单元格"""
        for merged_range in merged_ranges:
            if (merged_range.min_row <= cell.row <= merged_range.max_row and
                merged_range.min_col <= cell.column <= merged_range.max_col):
                # 如果是合并区域的主单元格
                if cell.row == merged_range.min_row and cell.column == merged_range.min_col:
                    return cell.value
                # 如果是合并区域的其他单元格，返回主单元格的值
                return merged_range.start_cell.value
        return cell.value if cell.value is not None else ""
    
    def _create_markdown_table_row(self, cells: List[str]) -> str:
        """创建Markdown表格行"""
        return "| " + " | ".join(cells) + " |\n"
    
    def _create_markdown_separator(self, num_columns: int) -> str:
        """创建Markdown表格分隔线"""
        return "| " + " | ".join(["---"] * num_columns) + " |\n"


class OnlineExcelParser(DocumentParser):
    """
    在线Excel文档解析器
    实现与远程服务器接口兼容的Excel文档解析功能
    简化输出结构，只保留原文件和转换后的MD文件
    """
    
    def __init__(self):
        super().__init__()
        self.local_parser = ExcelParser()
        logger.info("初始化在线Excel解析器 (简化版)")
    
    def parse(self, 
             file_bytes: bytes, 
             filename: str,
             output_dir: Path, 
             opts: Dict[str, Any]) -> Optional[Path]:
        """
        解析Excel文档并生成Markdown
        
        Args:
            file_bytes: 字节流形式的Excel文件内容
            filename: 原始文件名（带扩展名）
            output_dir: 结构化输出目录
            opts: 附加选项字典
        
        Returns:
            Path: 成功时返回生成的Markdown文件Path对象，失败时返回None
        """
        logger.info(f"开始解析Excel文档: {filename}")
        request_id = opts.get('request_id', Path(filename).stem)
        
        # 创建请求专属目录
        request_dir = output_dir / request_id
        auto_dir = request_dir / 'auto'
        os.makedirs(auto_dir, exist_ok=True)
        
        temp_path = None
        try:
            # 创建带正确后缀的临时文件
            suffix = Path(filename).suffix  # 保留原始扩展名(.xls/.xlsx)
            with tempfile.NamedTemporaryFile(
                mode='wb',
                suffix=suffix,
                delete=False
            ) as tmp_file:
                tmp_file.write(file_bytes)
                temp_path = Path(tmp_file.name)
            
            logger.info(f"临时文件已创建: {temp_path}")

            # 调用本地解析器
            md_path_str = self.local_parser.parse(
                str(temp_path),    # 本地解析器需要字符串路径
                str(auto_dir)      # 使用auto子目录
            )

            # 验证输出结果
            if md_path_str:
                md_path = Path(md_path_str)
                if md_path.exists() and md_path.stat().st_size > 0:
                    # 创建与请求ID同名的Markdown文件
                    final_md_path = auto_dir / f"{request_id}.md"
                    
                    # 如果生成的文件名与请求ID不同，重命名文件
                    if md_path.name != final_md_path.name:
                        shutil.copy2(md_path, final_md_path)
                        logger.info(f"已复制Markdown文件: {md_path} -> {final_md_path}")
                        
                        # 删除原始MD文件
                        md_path.unlink()
                        logger.debug(f"已删除原始Markdown文件: {md_path}")
                    
                    # 保存原始文件
                    orig_file_path = auto_dir / filename
                    with open(orig_file_path, 'wb') as f:
                        f.write(file_bytes)
                    logger.info(f"已保存原始文件: {orig_file_path}")
                    
                    logger.info(f"Excel解析完成，输出文件: {final_md_path}")
                    return final_md_path
            
            logger.error("Excel解析失败，未生成有效的Markdown文件")
            return None

        except Exception as e:
            logger.error(f"Excel解析异常: {str(e)}")
            raise RuntimeError(f"Excel解析失败: {str(e)}") from e
        finally:
            # 确保清理临时文件
            if temp_path and temp_path.exists():
                try:
                    temp_path.unlink()
                    logger.debug(f"临时文件已清理: {temp_path}")
                except Exception as cleanup_e:
                    logger.warning(f"临时文件清理失败: {str(cleanup_e)}")


if __name__ == "__main__":
    # 配置控制台日志
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # 测试本地解析
    input_path = "/path/to/test/spreadsheet.xlsx"
    output_path = "/path/to/output"
    
    if os.path.exists(input_path):
        excelparser = ExcelParser()
        result = excelparser.parse(input_path, output_path)
        print(f"解析结果: {result}")
    else:
        print(f"测试文件不存在: {input_path}")
