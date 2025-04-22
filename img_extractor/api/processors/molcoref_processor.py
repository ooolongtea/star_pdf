import json
import os
import re
import sys
import torch
import cv2
import numpy as np
from PIL import Image
import subprocess
from typing import Optional, Union, List, Dict, Any, Tuple
from pathlib import Path
from huggingface_hub import hf_hub_download

# Model imports
from molscribe import MolScribe
from rxnscribe import MolDetect, RxnScribe

class MolCorefProcessor:
    """分子共指解析处理器"""
    SUPPORTED_EXTENSIONS = {'.png', '.jpg', '.jpeg'}

    def __init__(self,
                 result_manager,
                 device: str = "cpu",
                 model_paths: Dict[str, str] = None):
        """
        初始化处理器

        参数:
            result_manager: 用于存储结果的ResultManager实例
            device: 计算设备('cpu'或'cuda:X')
            model_paths: 模型的可选路径(如果为None，则使用默认值)
        """
        self.result_manager = result_manager
        self.device = device

        self.model_paths = model_paths or {
            "molscribe": hf_hub_download('yujieq/MolScribe', 'swin_base_char_aux_1m.pth'),
            "molcoref": hf_hub_download("Ozymandias314/MolDetectCkpt", "coref_best_hf.ckpt"),
            "rxn": hf_hub_download("yujieq/RxnScribe", "pix2seq_reaction_full.ckpt")
        }

        # 初始化设备
        self._setup_device()

        # 加载模型
        self._load_models()

        # 标记模型加载状态
        self.models_loaded = True

    def _setup_device(self):
        """设置处理设备"""
        print(f"设置处理设备: {self.device}")

        if self.device.startswith('cuda'):
            if not torch.cuda.is_available():
                print("警告: CUDA不可用，回退到CPU")
                self.device = "cpu"
                self._device = torch.device("cpu")
            else:
                try:
                    # 解析设备ID
                    device_id = int(self.device.split(':')[1]) if ':' in self.device else 0

                    # 检查设备ID是否有效
                    if device_id < torch.cuda.device_count():
                        # 设置当前设备
                        torch.cuda.set_device(device_id)
                        self._device = torch.device(f"cuda:{device_id}")

                        # 打印设备信息
                        device_name = torch.cuda.get_device_name(device_id)
                        device_memory = torch.cuda.get_device_properties(device_id).total_memory / (1024**3)
                        print(f"成功设置CUDA设备 {device_id}: {device_name} (显存: {device_memory:.2f} GB)")
                    else:
                        print(f"警告: 设备ID {device_id} 超出范围 (0-{torch.cuda.device_count()-1})，回退到设备 0")
                        torch.cuda.set_device(0)
                        self._device = torch.device("cuda:0")
                        self.device = "cuda:0"
                except Exception as e:
                    print(f"设置CUDA设备失败: {e}，回退到CPU")
                    self.device = "cpu"
                    self._device = torch.device("cpu")
        else:
            # CPU设备
            self._device = torch.device("cpu")
            print("使用CPU设备")

    def _load_models(self):
        """加载所有所需的模型"""
        if not self.model_paths.get("molcoref"):
            default_path = os.path.expanduser("~/.cache/huggingface/hub/models--Ozymandias314--MolDetectCkpt/snapshots/7cf4ba5ffdae2aec35c8339693f1dafd70e0613c/coref_best_hf.ckpt")
            self.model_paths["molcoref"] = default_path

        try:
            # 加载分子检测模型
            print(f"加载分子检测模型，设备: {self.device}")
            try:
                self.model = MolDetect(
                    self.model_paths["molcoref"],
                    device=self._device,  # 使用初始化好的设备
                    coref=True  # 始终启用共指解析
                )
                print(f"  分子检测模型加载成功")
            except Exception as e:
                print(f"  分子检测模型加载失败: {e}")
                raise

            # 加载分子结构识别模型
            print(f"加载分子结构识别模型，设备: {self.device}")
            try:
                self.molscribe = MolScribe(
                    self.model_paths["molscribe"],
                    device=self.device
                )
                print(f"  分子结构识别模型加载成功")
            except Exception as e:
                print(f"  分子结构识别模型加载失败: {e}")
                raise

            # 加载反应模型
            print(f"加载反应模型，设备: {self.device}")
            try:
                self.rxnmodel = RxnScribe(
                    self.model_paths["rxn"],
                    device=self._device  # 使用初始化好的设备
                )
                print(f"  反应模型加载成功")
            except Exception as e:
                print(f"  反应模型加载失败: {e}")
                raise

            print("模型加载成功！---------------------------")
        except Exception as e:
            raise RuntimeError(f"加载模型失败：{str(e)}")

    def _should_skip(self, predictions: dict) -> bool:
        """检查是否需要跳过空结果"""
        bboxes = predictions.get('bboxes', [])
        corefs = predictions.get('corefs', [])
        return len(bboxes) == 0 and len(corefs) == 0

    def preprocess_image(self, image_path: Union[str, Path]) -> str:
        """
        图像预处理：去噪、锐化等

        参数:
            image_path: 原始图像路径

        返回:
            处理后的图像路径
        """
        img = cv2.imread(str(image_path), cv2.IMREAD_COLOR)
        if img is None:
            return str(image_path)  # 如果无法读取图像，返回原始路径

        # 转为灰度图
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        # 增强文本对比度
        binary = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)

        # 去噪
        denoised = cv2.fastNlMeansDenoising(binary, h=10)

        # 锐化图像
        kernel = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]])
        sharpened = cv2.filter2D(denoised, -1, kernel)

        # 保存处理后的图像
        processed_dir = Path(image_path).parent / "processed"
        processed_dir.mkdir(exist_ok=True)

        processed_path = processed_dir / Path(image_path).name
        cv2.imwrite(str(processed_path), sharpened)

        return str(processed_path)

    def process_image(self, image_path: Union[str, Path], output_dir: Optional[str] = None, need_mol: bool = True, need_rxn: bool = True) -> Optional[str]:
        """
        处理单个图像文件，可独立控制是否进行分子和反应预测

        参数:
            image_path: 图像文件的路径
            need_mol: 是否进行分子预测
            need_rxn: 是否进行反应预测

        返回:
            如果处理成功则返回图像ID，否则返回None
        """
        try:
            image_path = Path(image_path).resolve()

            # 验证输入
            if not image_path.exists():
                print(f"输入文件不存在：{image_path}")
                return None

            if not self._is_supported_image(image_path):
                print(f"不支持的文件类型：{image_path.suffix}")
                return None

            # 确保result_manager存在
            if self.result_manager is None:
                print(f"错误：result_manager未初始化")
                return None

            # 图像预处理
            processed_image_path = str(image_path)

            # 为此图像生成唯一ID
            image_id = f"{image_path.stem}"

            # 存储初始图像结果
            self.result_manager.store_image_result(image_id, str(image_path), {})

            # 执行预测（分子预测）
            if need_mol:
                try:
                    predictions = self.model.predict_image_file(processed_image_path, coref=True)
                    if not self._should_skip(predictions):
                        # 更新图像结果
                        self.result_manager.store_image_result(image_id, str(image_path), predictions)
                        # 处理分子和标识符
                        sheet_value = self._process_molecules_and_ids(predictions, image_path)
                        # 存储处理后的分子
                        for mol_data in sheet_value:
                            self.result_manager.store_molecule_result(image_id, mol_data)
                except Exception as e:
                    print(f"分子预测失败: {e}")

            # 执行预测（反应预测）
            if need_rxn:
                try:
                    rxn_predictions = self.rxnmodel.predict_image_file(processed_image_path, molscribe=True, ocr=True)
                    # 处理反应数据
                    reaction_data = self._process_reaction_data(rxn_predictions, image_path)
                    # 存储处理后的反应数据
                    for rxn_data in reaction_data:
                        self.result_manager.store_reaction_result(image_id, rxn_data)

                    # 如果需要，生成并存储可视化
                    if self.result_manager.output_config.get("visualization", True) and rxn_predictions:
                        vis_images = self.rxnmodel.draw_predictions(rxn_predictions, image_file=processed_image_path)
                        if vis_images:
                            self.result_manager.store_visualization(image_id, vis_images)
                except Exception as e:
                    print(f"反应预测失败: {e}")

            return image_id

        except Exception as e:
            print(f"图像处理异常: {image_path} - {str(e)}")
            return None

    def _is_supported_image(self, path: Path) -> bool:
        """检查文件类型是否受支持"""
        return path.suffix.lower() in self.SUPPORTED_EXTENSIONS

    def _process_reaction_data(self, rxn_predictions: list[dict], img_path: Path) -> List[dict]:
        """处理反应数据"""
        transformed_data = []

        # 只在需要时打开图片
        is_crop = True
        if is_crop:
            original_image = Image.open(img_path)
            W, H = original_image.size
            original_crop_image_dir = img_path.parent.parent / "image_reactions"
            original_crop_image_dir.mkdir(exist_ok=True)
            original_crop_image_path = original_crop_image_dir/ f"{img_path.stem}.png"
            original_image.save(original_crop_image_path)

        # 检查rxn_predictions是否为空或无效
        if not rxn_predictions or not isinstance(rxn_predictions, list):
            return transformed_data

        if is_crop:
            # 保存第一个反应物的图像
            reactant = rxn_predictions[0].get("reactants", [{}])
            if reactant and len(reactant) > 0:
                crop_img_index = len(reactant)
                for i in range(len(reactant)):
                    if reactant[i].get("smiles"):
                        reactant_bbox = rxn_predictions[0]["reactants"][i]['bbox']
                        reactant_crop = self._crop_image(original_image, reactant_bbox, W, H)
                        # crop_image_path = img_path.parent / f"{img_path.stem}_R0_S{i}.png"
                        crop_image_path = original_crop_image_dir / f"{img_path.stem}_R0_S{i}.png"
                        reactant_crop.save(crop_image_path)


        # 处理每个反应
        for i, reaction in enumerate(rxn_predictions):
            # 提取反应物 SMILES
            reactants_smiles = ""
            if reaction.get("reactants") and len(reaction["reactants"]) > 0:
                for r in reaction["reactants"] :
                    if r.get("smiles"):
                        reactants_smiles = ".".join(r.get("smiles", ""))

            # 提取产物 SMILES
            product_smiles = ""
            if i != 0:
                crop_img_index= 0
            if reaction.get("products") and len(reaction["products"]) > 0:
                for p in reaction["products"]:
                    if p.get("smiles"):
                        product_smiles = ".".join(p.get("smiles", ""))
                        # 裁剪所有产物图像
                        if is_crop:
                            product_bbox = p['bbox']
                            product_crop = self._crop_image(original_image, product_bbox, W, H)
                            crop_image_path = original_crop_image_dir / f"{img_path.stem}_R{i}_S{crop_img_index}.png"
                            product_crop.save(crop_image_path)
                            crop_img_index += 1

            # 提取条件
            condition = ""
            if reaction.get("conditions"):
                condition_texts = []
                for c in reaction["conditions"]:
                    if "text" in c and c["text"]:
                        # text 可能是字符串或列表
                        if isinstance(c["text"], list):
                            condition_texts.extend(filter(None, c["text"]))
                        else:
                            condition_texts.append(c["text"])
                condition = ", ".join(condition_texts)

            # 创建新的反应数据字典
            new_reaction = {
                "reactants_smiles": reactants_smiles,
                "product_smiles": product_smiles,
                "product_coref": "",  # 默认为空，可以后续与分子结果进行匹配
                "condition": condition,
                "row_img": "",
                "image_path": str(img_path),  # 保存原图路径，用于在Excel中显示
                "reaction_id": i, # 用来对应可视化图
            }
            transformed_data.append(new_reaction)

        return transformed_data

    def _process_molecules_and_ids(self, predictions: dict, img_path: Path) -> List[dict]:
        """处理分子及其标识符"""
        original_image = Image.open(img_path)
        W, H = original_image.size
        results = []
        processed_mol_indices = set()  # 记录已处理的分子索引
        mol_id = 0

        # 处理共指解析（分子-标识符对）
        for coref_idx, (mol_idx, idt_idx) in enumerate(predictions.get('corefs', [])):
            # 处理分子
            mol_bbox = predictions['bboxes'][mol_idx]['bbox']
            mol_crop = self._crop_image(original_image, mol_bbox, W, H)
            smiles = self.process_mol(mol_crop)

            # 处理标识符
            idt_bbox = predictions['bboxes'][idt_idx]['bbox']
            idt_crop = self._crop_image(original_image, idt_bbox, W, H)
            coref = self.ocr_image(idt_crop)

            results.append({
                "id": mol_id,
                "compound_smiles": smiles,
                "coref": coref,
                "compound_name": "",  # 添加空字段以符合Excel的格式
                "mol_idx": mol_idx,
                "idt_idx": idt_idx
            })
            processed_mol_indices.add(mol_idx)  # 记录已处理的分子
            mol_id += 1

        # 处理未被共指解析覆盖的[Mol]
        for mol_idx, bbox in enumerate(predictions['bboxes']):
            if bbox['category'] == '[Mol]' and mol_idx not in processed_mol_indices:
                mol_bbox = bbox['bbox']
                mol_crop = self._crop_image(original_image, mol_bbox, W, H)
                smiles = self.process_mol(mol_crop)

                results.append({
                    "id": mol_id,
                    "compound_smiles": smiles,
                    "coref": "",  # 未关联的标识符设为空
                    "compound_name": ""  # 添加空字段以符合Excel的格式
                })
                processed_mol_indices.add(mol_idx)
                mol_id += 1

        return results

    def _crop_image(self, image: Image, bbox: List[float], W: int, H: int) -> Image:
        """使用归一化坐标裁剪图像"""
        x_1 = bbox[0] * W
        y_1 = bbox[1] * H
        x_2 = min(W, bbox[2] * W)
        y_2 = min(H, bbox[3] * H)
        return image.crop((x_1, y_1, x_2, y_2))

    def process_mol(self, image: Image) -> str:
        """读取分子结构"""
        image_np = np.array(image)
        results = self.molscribe.predict_image(image_np)
        return results['smiles'] if results else ""

    def ocr_image(self, image: Image) -> str:
        """使用外部进程或直接运行图像OCR"""
        # 保存临时图像文件
        temp_image_path = "/tmp/temp_image.png"
        image.save(temp_image_path)

        # 获取当前脚本目录的路径
        script_dir = Path(__file__).parent
        script_path = script_dir / "ocr_runner.py"

        # 使用子进程调用PaddleOCR环境
        try:
            # 尝试找到安装了PaddleOCR的Python可执行文件
            python_paths = [
                os.environ.get("PADDLEOCR_PYTHON", ""),  # 如果设置了，使用环境变量
                os.path.expanduser("~/.conda/envs/zxy-paddleocr/bin/python"),  # 默认路径
                "python"  # 退回到系统Python
            ]

            # 使用第一个可用的Python路径
            python_executable = next((p for p in python_paths if p and os.path.exists(p)), "python")

            # 运行子进程
            result = subprocess.run(
                [python_executable, str(script_path), temp_image_path],
                capture_output=True,
                text=True,
                check=True,
                env={"PYTHONWARNINGFILTER": "ignore"}  # 忽略警告
            )

            # 使用正则表达式提取OCR结果
            pattern = r'paddleocr--------------:\s*\[([^\]]+)\]'
            match = re.search(pattern, result.stdout)

            if match:
                # 解析列表内容
                content = match.group(1).strip()
                if not content:  # 处理空列表
                    return ""

                items = [s.strip().strip("'") for s in content.split(",")]
                return items[0] if items else ""
            return ""

        except subprocess.CalledProcessError as e:
            print(f"OCR子进程失败：{e.stderr}", file=sys.stderr)
            return ""

    def write_to_excel(self, output_file, molecules_data=None, reactions_data=None):
        """
        将分子和反应数据写入Excel文件

        参数:
            output_file: 输出文件路径
            molecules_data: 分子数据列表 (可选，默认使用self.molecules)
            reactions_data: 反应数据列表 (可选，默认使用self.reactions)
        """
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        import os

        # 使用提供的数据或默认值
        molecules = molecules_data if molecules_data is not None else self.molecules
        reactions = reactions_data if reactions_data is not None else self.reactions

        # 确保目录存在
        os.makedirs(os.path.dirname(os.path.abspath(output_file)), exist_ok=True)

        # 创建工作簿和工作表
        wb = Workbook()

        # 分子工作表
        if molecules:
            # 删除默认sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            # 创建分子sheet
            ws_molecules = wb.create_sheet("分子结构")

            # 准备数据
            molecule_rows = []
            for mol in molecules:
                # 正则表达式识别编号
                mol_id = mol.get("id", "")

                # 图像信息
                image_path = mol.get("image_path", "")
                page_number = mol.get("page", "")

                # SMILES和InChI
                smiles = mol.get("smiles", "")
                inchi = mol.get("inchi", "")
                inchi_key = mol.get("inchi_key", "")

                # 可视化路径
                visualization_path = mol.get("visualization_path", "")

                # 共指信息
                coref_group = mol.get("coref_group", "")

                # 预测准确度
                confidence = mol.get("confidence", 0)
                if confidence:
                    confidence_str = f"{confidence:.2f}"
                else:
                    confidence_str = ""

                # 添加到行
                molecule_rows.append({
                    "ID": mol_id,
                    "页码": page_number,
                    "SMILES": smiles,
                    "InChI": inchi,
                    "InChI Key": inchi_key,
                    "置信度": confidence_str,
                    "共指组": coref_group,
                    "图像路径": image_path,
                    "可视化路径": visualization_path
                })

            # 创建DataFrame
            if molecule_rows:
                df_molecules = pd.DataFrame(molecule_rows)

                # 写入工作表
                for r_idx, row in enumerate(dataframe_to_rows(df_molecules, index=False, header=True)):
                    for c_idx, value in enumerate(row):
                        ws_molecules.cell(row=r_idx+1, column=c_idx+1, value=value)

                # 设置列宽
                for col in ws_molecules.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length

                    # 最小宽度15，最大宽度50
                    adjusted_width = max(15, min(max_length + 2, 50))
                    ws_molecules.column_dimensions[column].width = adjusted_width

                # 设置标题行格式
                header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                header_font = Font(bold=True)

                for col in range(1, df_molecules.shape[1] + 1):
                    cell = ws_molecules.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font

        # 反应工作表
        if reactions:
            ws_reactions = wb.create_sheet("反应图式")

            # 准备数据
            reaction_rows = []
            for rxn in reactions:
                # ID和图像信息
                rxn_id = rxn.get("id", "")
                image_path = rxn.get("image_path", "")
                page_number = rxn.get("page", "")

                # 反应SMILES
                rxn_smiles = rxn.get("reaction_smiles", "")

                # 反应物和产物
                reactants = rxn.get("reactants", [])
                products = rxn.get("products", [])

                # 反应物和产物SMILES列表
                reactant_smiles = "; ".join([r.get("smiles", "") for r in reactants])
                product_smiles = "; ".join([p.get("smiles", "") for p in products])

                # 可视化路径
                visualization_path = rxn.get("visualization_path", "")

                # 预测准确度
                confidence = rxn.get("confidence", 0)
                if confidence:
                    confidence_str = f"{confidence:.2f}"
                else:
                    confidence_str = ""

                # 添加到行
                reaction_rows.append({
                    "ID": rxn_id,
                    "页码": page_number,
                    "反应SMILES": rxn_smiles,
                    "反应物SMILES": reactant_smiles,
                    "产物SMILES": product_smiles,
                    "置信度": confidence_str,
                    "图像路径": image_path,
                    "可视化路径": visualization_path
                })

            # 创建DataFrame
            if reaction_rows:
                df_reactions = pd.DataFrame(reaction_rows)

                # 写入工作表
                for r_idx, row in enumerate(dataframe_to_rows(df_reactions, index=False, header=True)):
                    for c_idx, value in enumerate(row):
                        ws_reactions.cell(row=r_idx+1, column=c_idx+1, value=value)

                # 设置列宽
                for col in ws_reactions.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length

                    # 最小宽度15，最大宽度100
                    adjusted_width = max(15, min(max_length + 2, 100))
                    ws_reactions.column_dimensions[column].width = adjusted_width

                # 设置标题行格式
                header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                header_font = Font(bold=True)

                for col in range(1, df_reactions.shape[1] + 1):
                    cell = ws_reactions.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font

        # 检查是否要添加可视化链接
        # 安全检查：确保result_manager不为None并且有output_config属性
        has_visualization = False
        if self.result_manager is not None and hasattr(self.result_manager, 'output_config'):
            if self.result_manager.output_config.get("visualization", True):
                has_visualization = True

        # 如果有可视化文件，添加超链接
        if has_visualization:
            # 添加分子可视化链接
            if molecules and "分子结构" in wb.sheetnames:
                ws_molecules = wb["分子结构"]

                # 找到可视化路径列
                visual_col = None
                for col in range(1, ws_molecules.max_column + 1):
                    if ws_molecules.cell(row=1, column=col).value == "可视化路径":
                        visual_col = col
                        break

                if visual_col:
                    for row in range(2, ws_molecules.max_row + 1):
                        visual_path = ws_molecules.cell(row=row, column=visual_col).value
                        if visual_path and os.path.exists(visual_path):
                            # 创建相对路径
                            rel_path = os.path.relpath(visual_path, os.path.dirname(output_file))
                            # 设置超链接
                            cell = ws_molecules.cell(row=row, column=1)  # ID列
                            cell.hyperlink = rel_path
                            cell.font = Font(color="0563C1", underline="single")

            # 添加反应可视化链接
            if reactions and "反应图式" in wb.sheetnames:
                ws_reactions = wb["反应图式"]

                # 找到可视化路径列
                visual_col = None
                for col in range(1, ws_reactions.max_column + 1):
                    if ws_reactions.cell(row=1, column=col).value == "可视化路径":
                        visual_col = col
                        break

                if visual_col:
                    for row in range(2, ws_reactions.max_row + 1):
                        visual_path = ws_reactions.cell(row=row, column=visual_col).value
                        if visual_path and os.path.exists(visual_path):
                            # 创建相对路径
                            rel_path = os.path.relpath(visual_path, os.path.dirname(output_file))
                            # 设置超链接
                            cell = ws_reactions.cell(row=row, column=1)  # ID列
                            cell.hyperlink = rel_path
                            cell.font = Font(color="0563C1", underline="single")

        # 保存文件
        wb.save(output_file)
        print(f"Excel文件已保存到: {output_file}")